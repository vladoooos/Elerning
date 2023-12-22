import io
from datetime import datetime
from openpyxl import Workbook
from django.http import HttpResponse
from django.contrib import admin
from openpyxl.drawing.image import Image
import matplotlib
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
import matplotlib.pyplot as plt
from .models import CustomUser

matplotlib.use('Agg')  # Use Agg backend (for PNGs, not interactive)


class CustomUserAdmin(admin.ModelAdmin):
    actions = ['export_to_excel']

    def export_to_excel(self, request, queryset):
        # Create a new Excel workbook and add a worksheet.
        wb = Workbook()
        ws = wb.active

        # Write the header row in the worksheet.
        header_row = ['ID', 'Name', 'Surname', 'Email', 'Is Active', 'Is Staff', 'Registration Date']
        for col_num, column_title in enumerate(header_row, 1):
            ws.cell(row=1, column=col_num, value=column_title)

        # Write data rows in the worksheet.
        for row_num, user in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=user.id)
            ws.cell(row=row_num, column=2, value=user.name)
            ws.cell(row=row_num, column=3, value=user.surname)
            ws.cell(row=row_num, column=4, value=user.email)
            ws.cell(row=row_num, column=5, value=user.is_active)
            ws.cell(row=row_num, column=6, value=user.is_staff)

            # Ensure timezone is None before writing to Excel
            registration_date = user.registration_date.replace(tzinfo=None) if user.registration_date else None
            ws.cell(row=row_num, column=7, value=registration_date)

        # Create a pie chart of registrations by month
        labels = []
        data = []
        for month in range(1, 13):
            month_users = queryset.filter(registration_date__month=month)
            labels.append(datetime(1900, month, 1).strftime('%B'))
            data.append(month_users.count())

        fig, ax = plt.subplots(figsize=(15, 6))  # Adjust the size as needed
        ax.bar(labels, data, color='skyblue')
        ax.set_ylabel('Number of Registrations')
        ax.set_xlabel('Month')
        ax.set_title('Registrations by Month')

        chart_io = io.BytesIO()
        canvas = FigureCanvas(fig)
        canvas.print_png(chart_io)
        plt.close(fig)
        chart_io.seek(0)

        # Add the chart to the Excel workbook.
        img = Image(chart_io)
        ws.add_image(img, 'H5')

        # Create the HTTP response with Excel content.
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=users_export.xlsx'
        wb.save(response)

        return response

    export_to_excel.short_description = "Export selected users to Excel"


admin.site.register(CustomUser, CustomUserAdmin)
